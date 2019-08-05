import pymysql
import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from tkinter import Tk
from tkinter.filedialog import asksaveasfilename

Tk().withdraw()
filename = asksaveasfilename(filetypes=[("Excel files", "*.xlsx")])

path = filename

wb = Workbook()

ws = wb.active

db = pymysql.connect(host='localhost', user='root', password='ashy1256', db='bflow_data', charset='utf8')
cursor = db.cursor()

border = Border(
left=Side(border_style='thin', color='000000'),
right=Side(border_style='thin', color='000000'),
top=Side(border_style='thin', color='000000'),
bottom=Side(border_style='thin', color='000000'),
)

ws.fill = PatternFill(bgColor='ffffff', fill_type = 'solid')
ws.font = Font(size = 10)
ws.alignment = Alignment(horizontal = 'center', vertical ='center')

def noZerodiv(a, b):
    if a != None and b != None:
        return round(b/a, 2)
    else:
        return None
    

for i in range(1, 9):
    sql = f'''
    SELECT 
    date,
    min(date),
    max(date),
    sum(brich_total_amount),
    sum(brich_qty),
    sum(brich_margin),
    sum(gmarket_total_amount),
    sum(gmarket_qty),
    sum(gmarket_margin),
    sum(auction_total_amount),
    sum(auction_qty),
    sum(auction_margin),
    sum(11st_total_amount),
    sum(11st_qty),
    sum(11st_margin),
    sum(wemakeprice_total_amount),
    sum(wemakeprice_qty),
    sum(wemakeprice_margin),
    sum(interpark_total_amount),
    sum(interpark_qty),
    sum(interpark_margin),
    sum(coupang_total_amount),
    sum(coupang_qty),
    sum(coupang_margin),
    sum(ssg_total_amount),
    sum(ssg_qty),
    sum(ssg_margin),
    sum(g9_total_amount),
    sum(g9_qty),
    sum(g9_margin)
    FROM sellstatistics
    WHERE month = {i}
    GROUP BY date
    '''

    cursor.execute(sql)
    rows = cursor.fetchall()

    
    endRow = ws.max_row + 2
    no = 0 + endRow
    
    
    for row in rows:
        no += 1
        week = row[0]
        weekstr = datetime.datetime.strftime(row[1], '%Y-%m-%d') + "~" + datetime.datetime.strftime(row[2], '%Y-%m-%d')
        brich_total_amount = row[3]
        brich_qty = row[4]
        brich_CT = noZerodiv(brich_qty, brich_total_amount)
        brich_margin = row[5]
        brich_profit = noZerodiv(brich_total_amount, brich_margin)
        gmarket_total_amount = row[6]
        gmarket_qty = row[7]
        gmarket_CT = noZerodiv(gmarket_qty, gmarket_total_amount)
        gmarket_margin = row[8]
        gmarket_profit = noZerodiv(gmarket_total_amount, gmarket_margin)
        auction_total_amount = row[9]
        auction_qty = row[10]
        auction_CT = noZerodiv(auction_qty, auction_total_amount)
        auction_margin = row[11]
        auction_profit = noZerodiv(auction_total_amount, auction_margin)
        st_total_amount = row[12]
        st_qty = row[13]
        st_CT = noZerodiv(st_qty, st_total_amount)
        st_margin = row[14]
        st_profit = noZerodiv(st_total_amount, st_margin)
        wemakeprice_total_amount = row[15] 
        wemakeprice_qty = row[16] 
        wemakeprice_CT = noZerodiv(wemakeprice_qty, wemakeprice_total_amount)
        wemakeprice_margin = row[17]
        wemakeprice_profit = noZerodiv(wemakeprice_total_amount, wemakeprice_margin)
        interpark_total_amount = row[18]
        interpark_qty = row[19]
        interpark_CT = noZerodiv(interpark_qty, interpark_total_amount)
        interpark_margin = row[20]
        interpark_profit = noZerodiv(interpark_total_amount, interpark_margin)
        coupnag_total_amount = row[21]
        coupnag_qty = row[22]
        coupnag_CT = noZerodiv(coupnag_qty, coupnag_total_amount)
        coupnag_margin = row[23]
        coupnag_profit = noZerodiv(coupnag_total_amount, coupnag_margin)
        ssg_total_amount = row[24]
        ssg_qty = row[25]
        ssg_CT = noZerodiv(ssg_qty, ssg_total_amount)
        ssg_margin = row[26]
        ssg_profit = noZerodiv(ssg_total_amount, ssg_margin)
        g9_total_amount = row[27]
        g9_qty = row[28]
        g9_CT = noZerodiv(g9_qty, g9_total_amount)
        g9_margin = row[29]
        g9_profit = noZerodiv(g9_total_amount, g9_margin)

        ws.cell(row = 1, column = 1).value = '일자'
        ws.cell(row = 1, column = 2).value = '브리치 거래액'
        ws.cell(row = 1, column = 3).value = '브리치 판매수'
        ws.cell(row = 1, column = 4).value = '브리치 객단가'
        ws.cell(row = 1, column = 5).value = '브리치 마진'
        ws.cell(row = 1, column = 6).value = '브리치 수익율'
        ws.cell(row = 1, column = 7).value = '지마켓 거래액'
        ws.cell(row = 1, column = 8).value = '지마켓 판매수'
        ws.cell(row = 1, column = 9).value = '지마켓 객단가'
        ws.cell(row = 1, column = 10).value = '지마켓 마진'
        ws.cell(row = 1, column = 11).value = '지마켓 수익율'
        ws.cell(row = 1, column = 12).value = '옥션 거래액'
        ws.cell(row = 1, column = 13).value = '옥션 판매수'
        ws.cell(row = 1, column = 14).value = '옥션 객단가'
        ws.cell(row = 1, column = 15).value = '옥션 마진'
        ws.cell(row = 1, column = 16).value = '옥션 수익율'
        ws.cell(row = 1, column = 17).value = '11번가 거래액'
        ws.cell(row = 1, column = 18).value = '11번가 판매수'
        ws.cell(row = 1, column = 19).value = '11번가 객단가'
        ws.cell(row = 1, column = 20).value = '11번가 마진'
        ws.cell(row = 1, column = 21).value = '11번가 수익율'
        ws.cell(row = 1, column = 22).value = '위메프 거래액'
        ws.cell(row = 1, column = 23).value = '위메프 판매수'
        ws.cell(row = 1, column = 24).value = '위메프 객단가'
        ws.cell(row = 1, column = 25).value = '위메프 마진'
        ws.cell(row = 1, column = 26).value = '위메프 수익율'
        ws.cell(row = 1, column = 27).value = '인터파크 거래액'
        ws.cell(row = 1, column = 28).value = '인터파크 판매수'
        ws.cell(row = 1, column = 29).value = '인터파크 객단가'
        ws.cell(row = 1, column = 30).value = '인터파크 마진'
        ws.cell(row = 1, column = 31).value = '인터파크 수익율'
        ws.cell(row = 1, column = 32).value = '쿠팡 거래액'
        ws.cell(row = 1, column = 33).value = '쿠팡 판매수'
        ws.cell(row = 1, column = 34).value = '쿠팡 객단가'
        ws.cell(row = 1, column = 35).value = '쿠팡 마진'
        ws.cell(row = 1, column = 36).value = '쿠팡 수익율'
        ws.cell(row = 1, column = 37).value = 'SSG 거래액'
        ws.cell(row = 1, column = 38).value = 'SSG 판매수'
        ws.cell(row = 1, column = 39).value = 'SSG 객단가'
        ws.cell(row = 1, column = 40).value = 'SSG 마진'
        ws.cell(row = 1, column = 41).value = 'SSG 수익율'
        ws.cell(row = 1, column = 42).value = 'G9 거래액'
        ws.cell(row = 1, column = 43).value = 'G9 판매수'
        ws.cell(row = 1, column = 44).value = 'G9 객단가'
        ws.cell(row = 1, column = 45).value = 'G9 마진'
        ws.cell(row = 1, column = 46).value = 'G9 수익율'

        ws.cell(row = no, column = 1).value = week
        ws.cell(row = no, column = 1).border = border
        ws.cell(row = no, column = 2).value = brich_total_amount
        ws.cell(row = no, column = 2).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 2).border = border
        ws.cell(row = no, column = 3).value = brich_qty
        ws.cell(row = no, column = 3).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 3).border = border
        ws.cell(row = no, column = 4).value = brich_CT
        ws.cell(row = no, column = 4).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 4).border = border
        ws.cell(row = no, column = 5).value = brich_margin
        ws.cell(row = no, column = 5).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 5).border = border
        ws.cell(row = no, column = 6).value = brich_profit
        ws.cell(row = no, column = 6).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 6).border = border
        ws.cell(row = no, column = 7).value = gmarket_total_amount
        ws.cell(row = no, column = 7).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 7).border = border
        ws.cell(row = no, column = 8).value = gmarket_qty
        ws.cell(row = no, column = 8).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 8).border = border
        ws.cell(row = no, column = 9).value = gmarket_CT
        ws.cell(row = no, column = 9).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 9).border = border
        ws.cell(row = no, column = 10).value = gmarket_margin
        ws.cell(row = no, column = 10).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 10).border = border
        ws.cell(row = no, column = 11).value = gmarket_profit
        ws.cell(row = no, column = 11).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 11).border = border
        ws.cell(row = no, column = 12).value = auction_total_amount
        ws.cell(row = no, column = 12).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 12).border = border
        ws.cell(row = no, column = 13).value = auction_qty
        ws.cell(row = no, column = 13).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 13).border = border
        ws.cell(row = no, column = 14).value = auction_CT
        ws.cell(row = no, column = 14).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 14).border = border
        ws.cell(row = no, column = 15).value = auction_margin
        ws.cell(row = no, column = 15).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 15).border = border
        ws.cell(row = no, column = 16).value = auction_profit
        ws.cell(row = no, column = 16).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 16).border = border
        ws.cell(row = no, column = 17).value = st_total_amount
        ws.cell(row = no, column = 17).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 17).border = border
        ws.cell(row = no, column = 18).value = st_qty
        ws.cell(row = no, column = 18).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 18).border = border
        ws.cell(row = no, column = 19).value = st_CT
        ws.cell(row = no, column = 19).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 19).border = border
        ws.cell(row = no, column = 20).value = st_margin
        ws.cell(row = no, column = 20).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 20).border = border
        ws.cell(row = no, column = 21).value = st_profit
        ws.cell(row = no, column = 21).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 21).border = border
        ws.cell(row = no, column = 22).value = wemakeprice_total_amount
        ws.cell(row = no, column = 22).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 22).border = border
        ws.cell(row = no, column = 23).value = wemakeprice_qty
        ws.cell(row = no, column = 23).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 23).border = border
        ws.cell(row = no, column = 24).value = wemakeprice_CT
        ws.cell(row = no, column = 24).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 24).border = border
        ws.cell(row = no, column = 25).value = wemakeprice_margin
        ws.cell(row = no, column = 25).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 25).border = border
        ws.cell(row = no, column = 26).value = wemakeprice_profit
        ws.cell(row = no, column = 26).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 26).border = border
        ws.cell(row = no, column = 27).value = interpark_total_amount
        ws.cell(row = no, column = 27).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 27).border = border
        ws.cell(row = no, column = 28).value = interpark_qty
        ws.cell(row = no, column = 28).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 28).border = border
        ws.cell(row = no, column = 29).value = interpark_CT
        ws.cell(row = no, column = 29).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 29).border = border
        ws.cell(row = no, column = 30).value = interpark_margin
        ws.cell(row = no, column = 30).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 30).border = border
        ws.cell(row = no, column = 31).value = interpark_profit
        ws.cell(row = no, column = 31).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 31).border = border
        ws.cell(row = no, column = 32).value = coupnag_total_amount
        ws.cell(row = no, column = 32).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 32).border = border
        ws.cell(row = no, column = 33).value = coupnag_qty
        ws.cell(row = no, column = 33).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 33).border = border
        ws.cell(row = no, column = 34).value = coupnag_CT
        ws.cell(row = no, column = 34).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 34).border = border
        ws.cell(row = no, column = 35).value = coupnag_margin
        ws.cell(row = no, column = 35).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 35).border = border
        ws.cell(row = no, column = 36).value = coupnag_profit
        ws.cell(row = no, column = 36).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 36).border = border
        ws.cell(row = no, column = 37).value = ssg_total_amount
        ws.cell(row = no, column = 37).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 37).border = border
        ws.cell(row = no, column = 38).value = ssg_qty
        ws.cell(row = no, column = 38).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 38).border = border
        ws.cell(row = no, column = 39).value = ssg_CT
        ws.cell(row = no, column = 39).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 39).border = border
        ws.cell(row = no, column = 40).value = ssg_margin
        ws.cell(row = no, column = 40).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 40).border = border
        ws.cell(row = no, column = 41).value = ssg_profit
        ws.cell(row = no, column = 41).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 41).border = border
        ws.cell(row = no, column = 42).value = g9_total_amount
        ws.cell(row = no, column = 42).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 42).border = border
        ws.cell(row = no, column = 43).value = g9_qty
        ws.cell(row = no, column = 43).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 43).border = border
        ws.cell(row = no, column = 44).value = g9_CT
        ws.cell(row = no, column = 44).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 44).border = border
        ws.cell(row = no, column = 45).value = g9_margin
        ws.cell(row = no, column = 45).number_format = '#,##0;[red]-#,##0'
        ws.cell(row = no, column = 45).border = border
        ws.cell(row = no, column = 46).value = g9_profit
        ws.cell(row = no, column = 46).number_format = '0.0%;[red]-0.0%'
        ws.cell(row = no, column = 46).border = border

    
    lastRow = ws.max_row
    fristRow = 1 + lastRow - len(rows)
    nowRow = lastRow + 1
    #brich
    ws.cell(row = lastRow + 1, column = 2).value = f'=sum(b{fristRow}:b{lastRow})'
    ws.cell(row = lastRow + 1, column = 2).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 2).border = border
    ws.cell(row = lastRow + 1, column = 3).value = f'=sum(c{fristRow}:c{lastRow})'
    ws.cell(row = lastRow + 1, column = 3).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 3).border = border
    ws.cell(row = lastRow + 1, column = 4).value = f'=average(d{fristRow}:d{lastRow})'
    ws.cell(row = lastRow + 1, column = 4).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 4).border = border
    ws.cell(row = lastRow + 1, column = 5).value = f'=sum(e{fristRow}:e{lastRow})'
    ws.cell(row = lastRow + 1, column = 5).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 5).border = border
    ws.cell(row = lastRow + 1, column = 6).value = f'=e{nowRow}/b{nowRow}'
    ws.cell(row = lastRow + 1, column = 6).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 6).border = border
    #gamrket
    ws.cell(row = lastRow + 1, column = 7).value = f'=sum(g{fristRow}:g{lastRow})'
    ws.cell(row = lastRow + 1, column = 7).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 7).border = border
    ws.cell(row = lastRow + 1, column = 8).value = f'=sum(h{fristRow}:h{lastRow})'
    ws.cell(row = lastRow + 1, column = 8).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 8).border = border
    ws.cell(row = lastRow + 1, column = 9).value = f'=average(i{fristRow}:i{lastRow})'
    ws.cell(row = lastRow + 1, column = 9).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 9).border = border
    ws.cell(row = lastRow + 1, column = 10).value = f'=sum(j{fristRow}:j{lastRow})'
    ws.cell(row = lastRow + 1, column = 10).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 10).border = border
    ws.cell(row = lastRow + 1, column = 11).value = f'=j{nowRow}/g{nowRow}'
    ws.cell(row = lastRow + 1, column = 11).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 11).border = border
    #auction
    ws.cell(row = lastRow + 1, column = 12).value = f'=sum(l{fristRow}:l{lastRow})'
    ws.cell(row = lastRow + 1, column = 12).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 12).border = border
    ws.cell(row = lastRow + 1, column = 13).value = f'=sum(m{fristRow}:m{lastRow})'
    ws.cell(row = lastRow + 1, column = 13).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 13).border = border
    ws.cell(row = lastRow + 1, column = 14).value = f'=average(n{fristRow}:n{lastRow})'
    ws.cell(row = lastRow + 1, column = 14).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 14).border = border
    ws.cell(row = lastRow + 1, column = 15).value = f'=sum(o{fristRow}:o{lastRow})'
    ws.cell(row = lastRow + 1, column = 15).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 15).border = border
    ws.cell(row = lastRow + 1, column = 16).value = f'=o{nowRow}/l{nowRow}'
    ws.cell(row = lastRow + 1, column = 16).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 16).border = border
    #11st
    ws.cell(row = lastRow + 1, column = 17).value = f'=sum(q{fristRow}:q{lastRow})'
    ws.cell(row = lastRow + 1, column = 17).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 17).border = border
    ws.cell(row = lastRow + 1, column = 18).value = f'=sum(r{fristRow}:r{lastRow})'
    ws.cell(row = lastRow + 1, column = 18).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 18).border = border
    ws.cell(row = lastRow + 1, column = 19).value = f'=average(s{fristRow}:s{lastRow})'
    ws.cell(row = lastRow + 1, column = 19).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 19).border = border
    ws.cell(row = lastRow + 1, column = 20).value = f'=sum(t{fristRow}:t{lastRow})'
    ws.cell(row = lastRow + 1, column = 20).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 20).border = border
    ws.cell(row = lastRow + 1, column = 21).value = f'=t{nowRow}/q{nowRow}'
    ws.cell(row = lastRow + 1, column = 21).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 21).border = border
    #wemakeparice
    ws.cell(row = lastRow + 1, column = 22).value = f'=sum(v{fristRow}:v{lastRow})'
    ws.cell(row = lastRow + 1, column = 22).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 22).border = border
    ws.cell(row = lastRow + 1, column = 23).value = f'=sum(w{fristRow}:w{lastRow})'
    ws.cell(row = lastRow + 1, column = 23).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 23).border = border
    ws.cell(row = lastRow + 1, column = 24).value = f'=average(x{fristRow}:x{lastRow})'
    ws.cell(row = lastRow + 1, column = 24).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 24).border = border
    ws.cell(row = lastRow + 1, column = 25).value = f'=sum(y{fristRow}:y{lastRow})'
    ws.cell(row = lastRow + 1, column = 25).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 25).border = border
    ws.cell(row = lastRow + 1, column = 26).value = f'=y{nowRow}/v{nowRow}'
    ws.cell(row = lastRow + 1, column = 26).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 26).border = border
    #interpark
    ws.cell(row = lastRow + 1, column = 27).value = f'=sum(aa{fristRow}:aa{lastRow})'
    ws.cell(row = lastRow + 1, column = 27).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 27).border = border
    ws.cell(row = lastRow + 1, column = 28).value = f'=sum(ab{fristRow}:ab{lastRow})'
    ws.cell(row = lastRow + 1, column = 28).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 28).border = border
    ws.cell(row = lastRow + 1, column = 29).value = f'=average(ac{fristRow}:ac{lastRow})'
    ws.cell(row = lastRow + 1, column = 29).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 29).border = border
    ws.cell(row = lastRow + 1, column = 30).value = f'=sum(ad{fristRow}:ad{lastRow})'
    ws.cell(row = lastRow + 1, column = 30).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 30).border = border
    ws.cell(row = lastRow + 1, column = 31).value = f'=ad{nowRow}/aa{nowRow}'
    ws.cell(row = lastRow + 1, column = 31).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 31).border = border
    #coupang
    ws.cell(row = lastRow + 1, column = 32).value = f'=sum(af{fristRow}:af{lastRow})'
    ws.cell(row = lastRow + 1, column = 32).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 32).border = border
    ws.cell(row = lastRow + 1, column = 33).value = f'=sum(ag{fristRow}:ag{lastRow})'
    ws.cell(row = lastRow + 1, column = 33).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 33).border = border
    ws.cell(row = lastRow + 1, column = 34).value = f'=average(ah{fristRow}:ah{lastRow})'
    ws.cell(row = lastRow + 1, column = 34).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 34).border = border
    ws.cell(row = lastRow + 1, column = 35).value = f'=sum(ai{fristRow}:ai{lastRow})'
    ws.cell(row = lastRow + 1, column = 35).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 35).border = border
    ws.cell(row = lastRow + 1, column = 36).value = f'=ai{nowRow}/af{nowRow}'
    ws.cell(row = lastRow + 1, column = 36).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 36).border = border
    #ssg
    ws.cell(row = lastRow + 1, column = 37).value = f'=sum(ak{fristRow}:ak{lastRow})'
    ws.cell(row = lastRow + 1, column = 37).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 37).border = border
    ws.cell(row = lastRow + 1, column = 38).value = f'=sum(al{fristRow}:al{lastRow})'
    ws.cell(row = lastRow + 1, column = 38).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 38).border = border
    ws.cell(row = lastRow + 1, column = 39).value = f'=average(am{fristRow}:am{lastRow})'
    ws.cell(row = lastRow + 1, column = 39).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 39).border = border
    ws.cell(row = lastRow + 1, column = 40).value = f'=sum(an{fristRow}:an{lastRow})'
    ws.cell(row = lastRow + 1, column = 40).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 40).border = border
    ws.cell(row = lastRow + 1, column = 41).value = f'=an{nowRow}/ak{nowRow}'
    ws.cell(row = lastRow + 1, column = 41).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 41).border = border
    #g9
    ws.cell(row = lastRow + 1, column = 42).value = f'=sum(ap{fristRow}:ap{lastRow})'
    ws.cell(row = lastRow + 1, column = 42).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 42).border = border
    ws.cell(row = lastRow + 1, column = 43).value = f'=sum(aq{fristRow}:aq{lastRow})'
    ws.cell(row = lastRow + 1, column = 43).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 43).border = border
    ws.cell(row = lastRow + 1, column = 44).value = f'=average(ar{fristRow}:ar{lastRow})'
    ws.cell(row = lastRow + 1, column = 44).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 44).border = border
    ws.cell(row = lastRow + 1, column = 45).value = f'=sum(as{fristRow}:as{lastRow})'
    ws.cell(row = lastRow + 1, column = 45).number_format = '#,##0;[red]-#,##0'
    ws.cell(row = lastRow + 1, column = 45).border = border
    ws.cell(row = lastRow + 1, column = 46).value = f'=as{nowRow}/ap{nowRow}'
    ws.cell(row = lastRow + 1, column = 46).number_format = '0.0%;[red]-0.0%'
    ws.cell(row = lastRow + 1, column = 46).border = border


wb.save(path)



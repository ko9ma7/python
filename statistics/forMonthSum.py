import pymysql
from openpyxl import Workbook

wb = Workbook()

ws = wb.create_sheet()

db = pymysql.connect(host='localhost', user='root', password='ashy1256', db='bflow_data', charset='utf8')
cursor = db.cursor()

sql = '''SELECT month, channel, sum(total_amount) as total_amount,
sum(quantity) as quantity,
sum(calculate) as calaulate,
sum(calculate_channel) as channel_calaulate,
sum(margin) as margin
FROM sell
WHERE NOT order_state IN ('결제취소', '반품')
AND payment_at IS NOT NULL
GROUP BY month, channel'''

cursor.execute(sql)
rows = cursor.fetchall()

no = 1

for row in rows:
    no += 1
    month = row[0]
    channel = row[1]
    total_amount = row[2]
    quantity = row[3]
    ct = round(total_amount/quantity,0)
    calaulate = row[4]
    channel_calaulate = row[5]
    margin = row[6] or 0
    profit = round(margin/total_amount*100,2)
    ws.cell(row = 1, column = 1).value = '월'
    ws.cell(row = 1, column = 2).value = '채널'
    ws.cell(row = 1, column = 3).value = '총거래액'
    ws.cell(row = 1, column = 4).value = '판매수'
    ws.cell(row = 1, column = 5).value = '객단가'
    ws.cell(row = 1, column = 6).value = '정산금액'
    ws.cell(row = 1, column = 7).value = '채널정산 금액'
    ws.cell(row = 1, column = 8).value = '마진'
    ws.cell(row = 1, column = 9).value = '수익율'
    
    ws.cell(row = no, column = 1).value = month
    ws.cell(row = no, column = 2).value = channel
    ws.cell(row = no, column = 3).value = total_amount
    ws.cell(row = no, column = 4).value = quantity
    ws.cell(row = no, column = 5).value = ct
    ws.cell(row = no, column = 6).value = calaulate
    ws.cell(row = no, column = 7).value = channel_calaulate
    ws.cell(row = no, column = 8).value = margin
    ws.cell(row = no, column = 9).value = profit

wb.save('./text2.xlsx')




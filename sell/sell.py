import pymysql
import datetime
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

Tk().withdraw()
filename = askopenfilename()

path = filename

load_wb = load_workbook(path)

load_ws = load_wb['브리치 주문 내역']

db = pymysql.connect(host='localhost', user='root', password='ashy1256', db='bflow_data', charset='utf8')
cursor = db.cursor()


def replacedate(text):
    if text is None:
        return None
    else:
        text = text[0:10]
        return text


def replacenone(text):
    if text is None:
        return None
    else:
        text = str(text)
        return text


def replaceint(text):
    if text is None:
        return None
    else:
        text = int(text)
        return text


sql = '''INSERT INTO `bflow_data`.`sell` (
        product_order_number,
        order_number,
        payment_at,
        order_state,
        claim,
        provide_name,
        product_name,
        product_option,
        channel,
        product_number,
        product_amount,
        option_amount,
        seller_discount,
        quantity,
        total_amount,
        delivery_at,
        delivery_complete,
        order_complete_at,
        auto_complete_at,
        category_number,
        buyer_email,
        buyer_gender,
        buyer_age,
        crawler,
        week,
        month
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE payment_at = %s, order_state = %s, claim = %s,
         delivery_at = %s, delivery_complete = %s, order_complete_at =%s, auto_complete_at = %s, week = %s, month = %s
         '''

iter_rows = iter(load_ws.rows)
next(iter_rows)

for row in iter_rows:
    product_order_number = replacenone(row[0].value)
    order_number = replacenone(row[1].value)
    payment_at = replacedate(row[2].value)
    order_state = replacenone(row[3].value)
    claim = replacenone(row[4].value)
    provide_name = replacenone(row[5].value)
    product_name = replacenone(row[6].value)
    product_option = replacenone(row[7].value)
    channel = replacenone(row[8].value)
    product_number = replacenone(row[18].value)
    product_amount = replaceint(row[19].value)
    option_amount = replaceint(row[20].value)
    seller_discount = replaceint(row[21].value)
    quantity = replaceint(row[22].value)
    total_amount = replaceint(row[23].value)
    delivery_at = replacedate(row[24].value)
    delivery_complete = replacedate(row[25].value)
    order_complete_at = replacedate(row[26].value)
    auto_complete_at = replacedate(row[27].value)
    category_number = replacenone(row[39].value)
    buyer_email = replacenone(row[40].value)
    buyer_gender = replacenone(row[41].value)
    buyer_age = replacenone(row[42].value)
    crawler = replacenone(row[43].value)
    if payment_at is not None:
            monthStr = datetime.datetime.strptime(payment_at, '%Y-%m-%d')
            week = monthStr.isocalendar()[1]
            month = monthStr.month
    else:
            week = None
            month = None
        
    values = (
    product_order_number, order_number, payment_at, order_state, claim, provide_name, product_name, product_option,
    channel, product_number, product_amount, option_amount, seller_discount, quantity, total_amount, delivery_at,
    delivery_complete, order_complete_at, auto_complete_at, category_number, buyer_email, buyer_gender, buyer_age,
    crawler, week, month, payment_at, order_state, claim, delivery_at,
    delivery_complete, order_complete_at, auto_complete_at, week, month)

    print(sql, values)
    cursor.execute(sql, values)
    db.commit()

db.close()

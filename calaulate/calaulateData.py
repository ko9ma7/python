import pymysql
import datetime
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

Tk().withdraw()
filename = askopenfilename()

path = filename

load_wb = load_workbook(path)

load_ws = load_wb['브리치 외부채널 주문 목록']

db = pymysql.connect(host='localhost', user='root', password='ashy1256', db='bflow_data', charset='utf8')
cursor = db.cursor()


def replacedate(text):
    if text is None:
        return None
    else:
        text = text[0:10]
        return text


def replacenone(text):
    if text is None:
        return None
    else:
        text = str(text)
        return text


def replaceint(text):
    if text is None:
        return None
    else:
        text = int(text)
        return text


def replaceMustint(text):
    if text is None:
        return 0
    else:
        text = int(text)
        return text


sql = '''INSERT INTO `bflow_data`.`calculate` (
        product_order_number,
        order_number,
        channel_order_number,
        order_state,
        delivery_complete_at,
        provide_name,
        channel,
        quantity,
        order_amount,
        fees,
        calculate,
        channel_calculate,
        complete_at,
        matching_at,
        week,
        month,
        margin
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s )
        ON DUPLICATE KEY UPDATE order_state = %s,delivery_complete_at = %s, calculate = %s, channel_calculate = %s, margin = %s
         '''
iter_rows = iter(load_ws.rows)
next(iter_rows)

for row in iter_rows:
    product_order_number = replacenone(row[0].value)
    order_number = replacenone(row[1].value)
    channel_order_number = replacenone(row[2].value)
    order_state = replacenone(row[3].value)
    delivery_complete_at = replacedate(row[4].value)
    provide_name = replacenone(row[5].value)
    channel = replacenone(row[6].value)
    quantity = replaceint(row[7].value)
    order_amount = replaceint(row[8].value)
    fees = float(row[9].value)
    calculate = replaceMustint(row[10].value)
    channel_calculate = replaceMustint(row[11].value)
    complete_at = replacedate(row[12].value)
    matching_at = replacedate(row[13].value)
    monthStr = datetime.datetime.strptime(complete_at, '%Y-%m-%d')
    week = monthStr.isocalendar()[1]
    month = monthStr.month
    margin = channel_calculate - calculate

    values = (product_order_number,
              order_number,
              channel_order_number,
              order_state,
              delivery_complete_at,
              provide_name,
              channel,
              quantity,
              order_amount,
              fees,
              calculate,
              channel_calculate,
              complete_at,
              matching_at,
              week,
              month,
              margin,
              order_state,
              delivery_complete_at,
              calculate,
              channel_calculate,
              margin
              )

    print(sql, values)
    cursor.execute(sql, values)
    db.commit()

db.close()

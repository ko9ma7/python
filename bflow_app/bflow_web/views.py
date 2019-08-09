from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.db.models import Sum, F, Case, When, Q
from .models import Sell, Calculate
import openpyxl
import datetime


def replacedate(text):
    if text is None:
        return None
    else:
        text = text[0:10]
        return text


def replacenone(text):
    if text is None:
        return None
    else:
        text = str(text).replace(" ","")
        return text


def replaceint(text):
    if text is None:
        return None
    else:
        text = int(text)
        return text


def replaceMustint(text):
    if text is None:
        return 0
    else:
        text = int(text)
        return text


def sell_list(request):
    if request.method == "GET":
        sell_list = Sell.objects.filter(
                payment_at__isnull=False,
            ).filter(
                ~Q(order_state='결제취소'),
            ).exclude(
                # order_state = '결제취소',    
            ).values(
                'payment_at',
                'channel',
            ).annotate(
                total_amount=Sum('total_amount'),
                quantity=Sum('quantity'),
            ).annotate(
                ct=F('total_amount')/F('quantity')
            )
        print(str(sell_list.query))
        paginator = Paginator(sell_list, 20)
        page = request.GET.get('page')
        sell = paginator.get_page(page)
        return render(request, 'bflow_web/sell_list.html', {"sell": sell})
    else:
        return redirect('/sell/list')
        

def sell_create(request):
    if request.method == "POST":

        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        ws = wb.active
        print(ws)

        iter_rows = iter(ws.rows)
        next(iter_rows)

        for row in iter_rows:
            product_order_number = replacenone(row[0].value)
            order_number = replacenone(row[1].value)
            payment_at = replacedate(row[2].value)
            order_state = replacenone(row[3].value)
            claim = replacenone(row[4].value)
            provide_name = replacenone(row[5].value)
            product_name = replacenone(row[6].value)
            product_option = replacenone(row[7].value)
            channel = replacenone(row[8].value)
            product_number = replacenone(row[18].value)
            product_amount = replaceint(row[19].value)
            option_amount = replaceint(row[20].value)
            seller_discount = replaceint(row[21].value)
            quantity = replaceint(row[22].value)
            total_amount = replaceint(row[23].value)
            delivery_at = replacedate(row[24].value)
            delivery_complete = replacedate(row[25].value)
            order_complete_at = replacedate(row[26].value)
            auto_complete_at = replacedate(row[27].value)
            category_number = replacenone(row[39].value)
            buyer_email = replacenone(row[40].value)
            buyer_gender = replacenone(row[41].value)
            buyer_age = replacenone(row[42].value)
            crawler = replacenone(row[43].value)
            if payment_at is not None:
                    monthStr = datetime.datetime.strptime(payment_at, '%Y-%m-%d')
                    week = monthStr.isocalendar()[1]
                    month = monthStr.month
            else:
                    week = None
                    month = None
        

            sell = Sell.objects.update_or_create(
                product_order_number = product_order_number,
                
                defaults = {
                    'order_number' : order_number,
                    'provide_name' : provide_name,
                    'product_name' : product_name,
                    'product_option' : product_option,
                    'channel' : channel,
                    'product_number' : product_number,
                    'product_amount' : product_amount,
                    'option_amount' : option_amount,
                    'seller_discount' : seller_discount,
                    'quantity' : quantity,
                    'total_amount' : total_amount,
                    'category_number' : category_number,
                    'buyer_email' : buyer_email,
                    'buyer_gender' : buyer_gender,
                    'buyer_age' : buyer_age,
                    'crawler' : crawler,
                    'order_state':order_state,
                    'claim' : claim,
                    'payment_at' : payment_at,
                    'delivery_at' : delivery_at,
                    'delivery_complete' : delivery_complete,
                    'order_complete_at' : order_complete_at,
                    'auto_complete_at' : auto_complete_at,
                    'week' : week,
                    'month' : month,
                },
            )
        msg = f'{excel_file} : 업데이트 완료'

        return render(request, 'bflow_web/sell_list.html', {"msg": msg})
    else:
        return redirect('/sell/list')

def calculate_list(request):
    if request.method == "GET":
        sell_list = Sell.objects.filter(
                payment_at__isnull=False,
            ).filter(
                ~Q(order_state='결제취소'),
            ).exclude(
                # order_state = '결제취소',    
            ).values(
                'payment_at',
                'channel',
            ).annotate(
                total_amount=Sum('total_amount'),
                quantity=Sum('quantity'),
            ).annotate(
                ct=F('total_amount')/F('quantity')
            )
        print(str(sell_list.query))
        paginator = Paginator(sell_list, 20)
        page = request.GET.get('page')
        sell = paginator.get_page(page)
        return render(request, 'bflow_web/calculate_list.html', {"sell": sell})
    else:
        return redirect('/sell/list')        

def calculate_create(request):
    if request.method == "POST":

        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file)

        ws = wb.active
        print(ws)

        iter_rows = iter(ws.rows)
        next(iter_rows)

        for row in iter_rows:

            product_order_number = replacenone(row[0].value)
            order_number = replacenone(row[1].value)
            channel_order_number = replacenone(row[2].value)
            provide_name = replacenone(row[5].value)
            channel = replacenone(row[6].value)
            quantity = replaceint(row[7].value)
            order_amount = replaceint(row[8].value)
            fees = float(row[9].value)
            order_state = replacenone(row[3].value)
            delivery_complete_at= replacedate(row[4].value)
            calculate= replaceMustint(row[10].value)
            channel_calculate= replaceMustint(row[11].value)
            complete_at= replacedate(row[12].value)
            matching_at= replacedate(row[13].value)
            monthStr = datetime.datetime.strptime(complete_at, '%Y-%m-%d')
            week = monthStr.isocalendar()[1]
            month = monthStr.month
            # margin = channel_calculate - calculate


            obj, calculate = Calculate.objects.update_or_create(
                product_order_number = product_order_number,
                
                defaults = {
                    'order_number' : order_number,
                    'channel_order_number' : channel_order_number,
                    'provide_name' : provide_name,
                    'channel' : channel,
                    'quantity' : quantity,
                    'order_amount' : order_amount,
                    'fees' : fees,
                    'order_state': order_state,
                    'delivery_complete_at': delivery_complete_at,
                    'calculate': calculate,
                    'channel_calculate': channel_calculate,
                    'complete_at': complete_at,
                    'matching_at': matching_at,
                    'week' : week,
                    'month' : month,
                },
            )
        
        msg = f'{excel_file} : 업데이트 완료'

        return render(request, 'bflow_web/calculate_list.html', {"msg": msg})
    else:
        return redirect('/calculate/list')



# Create your views here.
